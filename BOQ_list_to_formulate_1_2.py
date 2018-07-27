#从清单中读取计量单位，从wbs表中读取明细工程量，然后在计量单中填写计算式。
#请保证wbs明细中b列是计量单g4拆分后的部位明细+清单号，例如'路基工程主线路基排水工程排水沟排水沟K26+200～K27+571'，c列是计量量；
#请保证boq表a列为编号，b列为名称，c列为单位；计量表为计量软件导出表。

import openpyxl
import pprint


#拆分g4部位明细，如果是单个部位则返回单个部位列表；如果是多个部位，则返回('路基工程主线路基排水工程排水沟排水沟K26+200～K27+571', 'K26+200～K27+571')列表
def word_clean(source_str):
    list0=source_str.replace('_x000D_','').replace(' ','').split('/') #按'/'分割后的部位列表
    list1=[]
    feature_list=[]
    feature=''
    if source_str.find('/')==-1:#如果不含'/'则直接返回原文本
        feature_list.append(source_str)
        return feature_list
    elif list0[-1].find(',')==-1:#如果含'/'但最后一项不含','则直接去'/'连接
        for word in list0:
            feature+=word
        feature_list.append(feature)
        return feature_list
    else:#如果含'/'且最后一项含','则返回多个计量部位与子部位配对的元组
        list1=list0[-1].split(',')
        for i in range(len(list0)-1):
            feature+=list0[i]
        for word in list1:
            feature_list.append(feature+str(word))
        return zip(feature_list,list1)


#获取wbs计量明细，部位明细+清单号为key（与word_clean子部位对应），计量量为value
def get_wbs_dic():
    wbs_loc=input('please enter wbs list route:')
    print('loading wbs list.........')
    wbs_wb=openpyxl.load_workbook(filename=wbs_loc)
    wbs_dic={}
    wbs_sh=wbs_wb.get_sheet_by_name('wbs')
    print('geting wbs dict.......')
    for i in range(2,7509):
        wbs_dic.setdefault(str(wbs_sh['b'+str(i)].value).replace('_x000D_','').replace(' ','')\
        .replace('/',''),wbs_sh['c'+str(i)].value)
    return wbs_dic


#获取boq，编号为key，单位为value[0],名称为value[1]
def get_boq_dic():
    boq_loc=input('please enter boq list route:')
    print('loading boq list.........')
    boq_wb=openpyxl.load_workbook(filename=boq_loc)
    boq_dic={}
    boq_sh=boq_wb.get_sheet_by_name('boq')
    print('geting boq dic........')
    for i in range(2,200):
        if str(boq_sh['c'+str(i)].value)=='总额':
            boq_dic.setdefault(boq_sh['a'+str(i)].value,['元','总则'])
        else:
            boq_dic.setdefault(boq_sh['a'+str(i)].value,[boq_sh['c'+str(i)].value,boq_sh['b'+str(i)].value])
    return boq_dic


#主程序部分
def cost_cal():
    wrong_list=[]#返回处理失败的表格列表
    #获取计量单路径并初始化计算式为空
    cost_loc=input('please enter cost list route:')
    print('loading and innitializing cost_list.....')
    cost_wb=openpyxl.load_workbook(filename=cost_loc)
    for sh in cost_wb:
        sh['a10']=''
    wbs_dic=get_wbs_dic()
    boq_dic=get_boq_dic()
    for sh in cost_wb:
        print('processing %s'%(sh.title))
        try:
            sh['e5']=boq_dic[sh['b4'].value][1]
            sh['a10']='本期计量：%s:'%(sh['e5'].value)
            unit=boq_dic[sh['b4'].value][0]
            if isinstance(word_clean(sh['g4'].value),list):
                sh['a10']=sh['a10'].value+str(sh['e7'].value)+str(unit)+';'+'合计:'+str(sh['e7'].value)+str(unit)
            else:
                for feature,word in word_clean(sh['g4'].value):
                    sh['a10']=sh['a10'].value+str(word)+':'+str(wbs_dic[str(feature)+sh['b4'].value])+unit+','
                sh['a10']=str(sh['a10'].value)+'合计：'
                for feature,word in word_clean(sh['g4'].value):
                    sh['a10']=str(sh['a10'].value)+str(wbs_dic[str(feature)+str(sh['b4'].value)])+'+'
                sh['a10']=str(sh['a10'].value)[:-1]+'='+str(sh['e7'].value)+unit
        except:
            print('something wrong in %s'%(sh.title))
            if sh['b4'].value in ['203-1-a','203-1-b','204-1-a','204-1-b']:
                wrong_list.append(str(sh.title)+':土石方计量请自行填写')
                sh['a10']='尝试失败,土石方计量请自行填写'
            else:
                wrong_list.append(sh.title)
                sh['a10']='尝试失败,且非土石方计量'
                #raise
    print('saving result...........')
    cost_wb.save(cost_loc)
    print('以下%s个表格发生错误:'%(len(wrong_list)))
    for wrong_title in wrong_list:
        pprint.pprint(wrong_title)
    print('finished')


if __name__=='__main__':
    cost_cal()
