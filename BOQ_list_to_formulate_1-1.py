def del_lastchar(char):
    de_str=''
    for word in list(char)[:-1]:
        de_str+=word
    return de_str
def word_clean(source_str):#如果是单个部位则返回单个部位列表；如果是多个部位，则返回多个部位与其子部位配对的元组
    list=source_str.replace('_x000D_\n','').replace(' ','').split('/') #按'/'分割后的部位列表
    list1=[]#接受list最后一项按','分割后的列表
    feature_list=[]#接受处理完毕的部位列表，每个部位对应wbs表一行
    feature=''#feature处理初始文本
    if source_str.find('/')==-1:#如果不含'/'则直接返回原文本
        feature_list.append(source_str)
        return feature_list
    elif list[len(list)-1].find(',')==-1:#如果含'/'但最后一项不含','则直接去'/'连接
        for word in list:
            feature+=word
        feature_list.append(feature)
        return feature_list
    else:#如果含'/'且最后一项含','则返回多个计量部位与子部位配对的元组
        list1=list[len(list)-1].split(',')
        for i in range(len(list)-1):
            feature+=list[i]
        for word in list1:
            feature_list.append(feature+str(word))
        return zip(feature_list,list1)
import openpyxl
import pprint
cost_loc=input('please enter cost list route:')
print('loading and innitializing cost_list.....')
cost_wb=openpyxl.load_workbook(filename=cost_loc)
for sh in cost_wb:
    sh['a10']=''
wbs_loc=input('please enter wbs list route:')
print('loading wbs list.........')
wbs_wb=openpyxl.load_workbook(filename=wbs_loc)
boq_loc=input('please enter boq list route:')
print('loading boq list.........')
boq_wb=openpyxl.load_workbook(filename=boq_loc)
wbs_dic={}#wbs列表字典，用于读取计量工程量
boq_dic={}#BOQ列表字典，用于读取清单单位
wbs_sh=wbs_wb.get_sheet_by_name('wbs')
boq_sh=boq_wb.get_sheet_by_name('boq')
wrong_list=[]#返回处理失败的表格列表
print('geting wbs dict.......')
for i in range(2,7000):#获取本次计量wbs列表保存为字典
    wbs_dic.setdefault(wbs_sh['b'+str(i)].value,wbs_sh['c'+str(i)].value)
print('geting boq dic........')    
for i in range(2,200):#获取清单编号对应单位保存为字典 
    if str(boq_sh['c'+str(i)].value)=='总额':
        boq_dic.setdefault(boq_sh['a'+str(i)].value,['元','总则'])
    else:
        boq_dic.setdefault(boq_sh['a'+str(i)].value,[boq_sh['c'+str(i)].value,boq_sh['b'+str(i)].value])   
for sh in cost_wb:
    print('processing %s'%(sh.title))
    try:
        sh['e5']=boq_dic[sh['b4'].value][1]
        sh['a10']='本期计量：%s:'%(sh['e5'].value)
        if isinstance(word_clean(sh['g4'].value),list):
            sh['a10']=sh['a10'].value+str(wbs_dic[str(word_clean(sh['g4'].value)[0])+sh['b4'].value])+str(boq_dic[sh['b4'].value][0])+';'+'合计:'+str(sh['e7'].value)+str(boq_dic[sh['b4'].value][0])
        else:
            for feature,word in word_clean(sh['g4'].value):
                sh['a10']=sh['a10'].value+str(word)+':'+str(wbs_dic[str(feature)+sh['b4'].value])+boq_dic[sh['b4'].value][0]+','
            sh['a10']=str(sh['a10'].value)+'合计：'
            for feature,word in word_clean(sh['g4'].value):
                sh['a10']=str(sh['a10'].value)+str(wbs_dic[str(feature)+str(sh['b4'].value)])+'+'
            sh['a10']=del_lastchar(sh['a10'].value)+'='+str(sh['e7'].value)+boq_dic[sh['b4'].value][0]
    except:
        print('something wrong in %s'%(sh.title))
        if sh['b4'].value in ['203-1-a','203-1-b','204-1-a','204-1-b']:
            wrong_list.append(str(sh.title)+':土石方计量请自行填写')
            sh['a10']='尝试失败,土石方计量请自行填写'
        #elif sh.title=='第24页':#错误排除页代码，以便在else中raise出想看错误的行
            #print('24')
        else:
            wrong_list.append(sh.title)
            sh['a10']='尝试失败,且非土石方计量'
            #raise
print('saving result...........')
cost_wb.save(cost_loc)
print('以下%s个表格发生错误:'%(len(wrong_list)))
for wrong_title in wrong_list:
    pprint.pprint(wrong_title)
print('finished')
