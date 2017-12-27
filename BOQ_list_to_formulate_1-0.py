<<<<<<< HEAD
def word_clean(source_str):
    list=source_str.split('/')  #接受按'/'分割后的部位列表
    list1=[]#接受list最后一项按','分割后的列表
    feature_list=[]#接受处理完毕的部位列表，每个部位对应wbs表一行
    feature=''#feature处理初始文本
    if source_str.find('/')==-1:#如果不含'/'则直接返回原文本
        feature_list.append(source_str)
        return feature_list
    elif list[len(list)-1].find(',')==-1:#如果含'/'但最后一项不含','则直接去'/'连接
        for word in list:
            feature+=word
        feature_list.append(feature)
        return feature_list
    else:#如果含'/'且最后一项含','则返回多个计量部位的列表
        list1=list[len(list)-1].split(',')
        for i in range(len(list)-1):
            feature+=list[i]
        for word in list1:
            feature_list.append(feature+str(word))
        return feature_list
import openpyxl
cost_wb=openpyxl.load_workbook(filename=r'c:\python tem\cost.xlsx')
wbs_wb=openpyxl.load_workbook(filename=r'c:\python tem\wbs.xlsx')
wbs_dic={}#wbs列表字典
wbs_sh=wbs_wb.get_sheet_by_name('wbs')
for i in range(2,7000):
    wbs_dic.setdefault(wbs_sh['b'+str(i)].value,wbs_sh['c'+str(i)].value)#获取本次计量wbs列表保存为字典
for sh in cost_wb:
    sh['a10']='本期计量：%s:'%(sh['e5'].value)
    for feature in word_clean(sh['g4'].value):
        sh['a10']=str(sh['a10'].value)+str(feature)+':'+str(wbs_dic[str(feature)+str(sh['b4'].value)])+','
    sh['a10']=str(sh['a10'].value)+'合计：'+str(sh['e7'].value)
=======
def word_clean(source_str):
    list=source_str.split('/')  #接受按'/'分割后的部位列表
    list1=[]#接受list最后一项按','分割后的列表
    feature_list=[]#接受处理完毕的部位列表，每个部位对应wbs表一行
    feature=''#feature处理初始文本
    if source_str.find('/')==-1:#如果不含'/'则直接返回原文本
        feature_list.append(source_str)
        return feature_list
    elif list[len(list)-1].find(',')==-1:#如果含'/'但最后一项不含','则直接去'/'连接
        for word in list:
            feature+=word
        feature_list.append(feature)
        return feature_list
    else:#如果含'/'且最后一项含','则返回多个计量部位的列表
        list1=list[len(list)-1].split(',')
        for i in range(len(list)-1):
            feature+=list[i]
        for word in list1:
            feature_list.append(feature+str(word))
        return feature_list
import openpyxl
cost_wb=openpyxl.load_workbook(filename=r'c:\python tem\cost.xlsx')
wbs_wb=openpyxl.load_workbook(filename=r'c:\python tem\wbs.xlsx')
wbs_dic={}#wbs列表字典
wbs_sh=wbs_wb.get_sheet_by_name('wbs')
for i in range(2,7000):
    wbs_dic.setdefault(wbs_sh['b'+str(i)].value,wbs_sh['c'+str(i)].value)#获取本次计量wbs列表保存为字典
for sh in cost_wb:
    sh['a10']='本期计量：%s:'%(sh['e5'].value)
    for feature in word_clean(sh['g4'].value):
        sh['a10']=str(sh['a10'].value)+str(feature)+':'+str(wbs_dic[str(feature)+str(sh['b4'].value)])+','
    sh['a10']=str(sh['a10'].value)+'合计：'+str(sh['e7'].value)
>>>>>>> fee347fd13a570677259dcba9a53ef747de71a1c
cost_wb.save(r'c:\python tem\cost.xlsx')